import openpyxl
workbook = openpyxl.load_workbook(f"E:/MOOC/MOOC_Improvement_Based_on_Aspect-based_Sentiment_Analysis_Driven_by_Large_Language_Models/aspects3.xlsx")
sheet = workbook.active
criteria = {}
criteria_and_explains = {}

last_primary_indicator = ''
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    primary_indicator, secondary_indicator, explanation = row[0], row[1], row[2]
    criteria_and_explains[secondary_indicator] = explanation
    if primary_indicator:
        criteria[primary_indicator] = [secondary_indicator]
        last_primary_indicator = primary_indicator
    else:
        criteria[last_primary_indicator].append(secondary_indicator)
# print("Criteria:", criteria)
# print("Criteria and Explains:", criteria_and_explains)



criteria2score = {i:{j:{'正面':0, '负面':0, '不相关':0} for j in criteria[i]} for i in criteria.keys()}

attribute_list = []
for i in criteria.keys():
  for j in criteria[i]:
    attribute_list.append(j)
# print(len(attribute_list))
#

import os
from langchain_openai import ChatOpenAI

from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain
import asyncio

api_key = os.getenv("ZHIPU_API_KEY")
llm = ChatOpenAI(
        model="glm-4-flash",
        openai_api_key=api_key,  # 请填写您自己的APIKey
        openai_api_base="https://open.bigmodel.cn/api/paas/v4/"
    )


import xlwt
import xlrd

dataname = '人工智能5'
wb = xlrd.open_workbook(f"E:/MOOC/MOOC_Improvement_Based_on_Aspect-based_Sentiment_Analysis_Driven_by_Large_Language_Models/data_after_preprocessing/{dataname}.xls")
sheet = wb.sheet_by_index(0)  # 通过索引的方式获取到某一个sheet，现在是获取的第一个sheet页，也可以通过sheet的名称进行获取，sheet_by_name('sheet名称')
rows = sheet.nrows  # 获取sheet页的行数，一共有几行
names = []
reviews = []
times = []
likes = []
class_order = []
rankings = []
for i in range(1, rows):
    names.append(str(sheet.cell(i, 0).value).strip())
    reviews.append(str(sheet.cell(i, 1).value).strip())
    times.append(str(sheet.cell(i, 2).value).strip())
    likes.append(str(sheet.cell(i, 3).value).strip())
    class_order.append(str(sheet.cell(i, 4).value).strip())
    rankings.append(str(sheet.cell(i, 5).value).strip())


criteria_for_each_review = [{j:{k:'0' for k in criteria[j]} for j in criteria.keys()} for _ in range(len(reviews))]

template = '''
    帮我分析课程评论"{review}"中关于方面"{aspect}"的情感，
    其中关于方面“{aspect}”情感的评判标准为“{explain}”。
    如果是肯定的，回复1
    如果是否定的，回复-1
    如果评论与“{aspect}”无关，回复0。即评论中不存在与“{aspect}”的近似表达，回复0。
    请不要过度延伸与解读评论内容，评论内容明确表达了方面“{aspect}”才能回复1或者-1。例如，"非常好，很实用的课程。"这条评论明显相关的方面只有"课程内容的实用性"，在课程内容的实用性方面回复1，其他方面回复都应该是0。
    即评论与“{aspect}”的相关标准是，评论中出现“{aspect}”的近似表达。
    即最后的回复是1，-1，0三者之一。
    只回复数字，不要加入你的分析。
    请一步步思考后，再给出最后的回复，务必确保，评论的确提到了“{aspect}”并且关于评论关于它是肯定的态度，才能回复1，评论的确提到了“{aspect}”并且关于评论关于它是否定的态度，才能回复-1。其他一切情况回复0。
    '''


prompt = PromptTemplate.from_template(template)
# print(prompt1)

chain = LLMChain(llm=llm, prompt=prompt)


book = xlwt.Workbook(encoding='utf-8')
sheet1 = book.add_sheet('test', cell_overwrite_ok=True)
sheet1.write(0, 0, '一级指标')
sheet1.write(0, 1, '二级指标')
sheet1.write(0, 2, '数量')
sheet1.write(0, 3, '数量')
sheet1.write(1, 2, '正面')
sheet1.write(1, 3, '负面')


book2 = xlwt.Workbook(encoding='utf-8')
sheet2 = book2.add_sheet('test', cell_overwrite_ok=True)
sheet2.write(0, 0, '用户昵称')
sheet2.write(0, 1, '评论内容')
sheet2.write(0, 2, '评论时间')
sheet2.write(0, 3, '点赞数')
sheet2.write(0, 4, '第几次课程')
sheet2.write(0, 5, '评分')


async def asy_chain_tool(chain, review, aspects, explains, semaphore):
    async with semaphore:
        input_list = [{'review': review, 'aspect': i, 'explain':explains[i]} for i in aspects]
        resp = await chain.aapply(input_list)
        result = {aspects[i]: resp[i]['text'] for i in range(len(aspects))}
        print(result)
        return result


async def asy_chain():
    semaphore = asyncio.Semaphore(5)
    tasks = [asy_chain_tool(chain, reviews[i], attribute_list, criteria_and_explains, semaphore) for i in range(len(reviews))]
    results_all = await asyncio.gather(*tasks)
    return results_all


results_all = asyncio.run(asy_chain())
print(results_all[:10])


for i in range(len(reviews)):
  for j in criteria2score.keys():
    for k in criteria2score[j].keys():
      if results_all[i][k] == '1':
        criteria2score[j][k]['正面'] += 1
        criteria_for_each_review[i][j][k] = '1'
      elif results_all[i][k] == '0':
        criteria2score[j][k]['不相关'] += 1
        criteria_for_each_review[i][j][k] = '0'
      elif results_all[i][k] == '-1':
        criteria2score[j][k]['负面'] += 1
        criteria_for_each_review[i][j][k] = '-1'
      # else:
      #   criteria_for_each_review[i][j][k] = '2'

  row = 2
  for ii in criteria2score.keys():
    sheet1.write(row, 0, ii)
    for jj in criteria2score[ii].keys():
      sheet1.write(row, 1, jj)
      sheet1.write(row, 2, criteria2score[ii][jj]['正面'])
      sheet1.write(row, 3, criteria2score[ii][jj]['负面'])
      # sheet1.write(row, 4, criteria2score[ii][jj]['无明显情感倾向'])
      row += 1
  book.save(
    f"E:/MOOC/MOOC_Improvement_Based_on_Aspect-based_Sentiment_Analysis_Driven_by_Large_Language_Models/results{dataname}.xls")
  sheet2.write(i + 2, 0, names[i])
  sheet2.write(i + 2, 1, reviews[i])
  sheet2.write(i + 2, 2, times[i])
  sheet2.write(i + 2, 3, likes[i])
  sheet2.write(i + 2, 4, class_order[i])
  sheet2.write(i + 2, 5, rankings[i])

  column = 6
  for ii in criteria2score.keys():
    sheet2.write(0, column, ii)
    for jj in criteria2score[ii].keys():
      sheet2.write(1, column, jj)
      sheet2.write(i + 2, column, criteria_for_each_review[i][ii][jj])
      column += 1
  book2.save(
    f"E:/MOOC/MOOC_Improvement_Based_on_Aspect-based_Sentiment_Analysis_Driven_by_Large_Language_Models/results/{dataname}_attributes_for_each_review.xls")

